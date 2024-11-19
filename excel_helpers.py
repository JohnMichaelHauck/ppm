from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import company_constants as cc
import product_variable_ranges as pvr

class ExcelHelpers():
    
    def row_from_label(self, sheet, start_row, label):
        for row in range(start_row, sheet.max_row + 1):
            cell_value = sheet.cell(row=row, column=1).value
            if cell_value == label:
                return row
        return 0

    def range_from_label(self, sheet, start_row, label):
        for row in range(start_row, sheet.max_row + 1):
            cell_value = sheet.cell(row=row, column=1).value
            if cell_value == None:
                return None
            if cell_value == label:
                return [
                    sheet.cell(row=row, column=2).value,
                    sheet.cell(row=row, column=3).value,
                    sheet.cell(row=row, column=4).value
                ]

    def read_excel_data(self, file_path):
        # Load the workbook and select the sheets
        workbook = load_workbook(file_path, data_only=True, read_only=True)
        company_constants_sheet = workbook['Company Constants']

        # Initialize the CompanyConstants instance
        company_constants = cc.CompanyConstants(
            market_return = company_constants_sheet.cell(row=2, column=2).value,
            yearly_development_fte_cost_pv = company_constants_sheet.cell(row=3, column=2).value,
            maximum_development_ftes = company_constants_sheet.cell(row=4, column=2).value,
            development_cost_trend = company_constants_sheet.cell(row=5, column=2).value,
            product_cost_trend = company_constants_sheet.cell(row=6, column=2).value,
            product_price_trend = company_constants_sheet.cell(row=7, column=2).value)

        mix_variables_ranges = []
        mix_sheet = workbook['Mix']
        product_variables_ranges = pvr.ProductVariablesRanges()

        for mix_sheet_row in range(2, mix_sheet.max_row + 1):
            pvr_sheet_name = mix_sheet.cell(row=mix_sheet_row, column=1).value
            if( pvr_sheet_name == None ):
                break
            pvr_name = mix_sheet.cell(row=mix_sheet_row, column=2).value
            pvr_type = mix_sheet.cell(row=mix_sheet_row, column=3).value
            pvr_exclude = mix_sheet.cell(row=mix_sheet_row, column=4).value
            pvr_sheet = workbook[pvr_sheet_name]
            pvr_sheet_row = self.row_from_label(pvr_sheet, 1, pvr_name)
            if( pvr_sheet_row == 0):
                continue

            if( pvr_type == "Product"):
                product_variables_ranges = pvr.ProductVariablesRanges(
                    name = pvr_name,
                    type = pvr_type,
                    years_of_development_growth = self.range_from_label(pvr_sheet, pvr_sheet_row, "Years of Development Growth"),
                    years_of_development_maturity = self.range_from_label(pvr_sheet, pvr_sheet_row, "Years of Development Maturity"),
                    years_of_development_decline = self.range_from_label(pvr_sheet, pvr_sheet_row, "Years of Development Decline"),
                    years_of_pilot = self.range_from_label(pvr_sheet, pvr_sheet_row, "Years of Pilot"),
                    years_of_sales_growth = self.range_from_label(pvr_sheet, pvr_sheet_row, "Years of Sales Growth"),
                    years_of_sales_maturity = self.range_from_label(pvr_sheet, pvr_sheet_row, "Years of Sales Maturity"),
                    years_of_sales_decline = self.range_from_label(pvr_sheet, pvr_sheet_row, "Years of Sales Decline"),
                    development_ftes = self.range_from_label(pvr_sheet, pvr_sheet_row, "Development FTEs"),
                    maintenance_ftes = self.range_from_label(pvr_sheet, pvr_sheet_row, "Maintenance FTEs"),
                    years_of_maintenance = self.range_from_label(pvr_sheet, pvr_sheet_row, "Years of Maintenance"),
                    unit_cost_pv = self.range_from_label(pvr_sheet, pvr_sheet_row, "Unit Cost"),
                    unit_margin = self.range_from_label(pvr_sheet, pvr_sheet_row, "Unit Margin"),
                    sga_factor = self.range_from_label(pvr_sheet, pvr_sheet_row, "SG&A"),
                    yearly_unit_sales = self.range_from_label(pvr_sheet, pvr_sheet_row, "Yearly Unit Sales"),
                    yearly_unit_consumable_sales = self.range_from_label(pvr_sheet, pvr_sheet_row, "Yearly Unit Consumable Sales"),
                    years_of_consumable_sales = self.range_from_label(pvr_sheet, pvr_sheet_row, "Years of Consumable Sales"),
                    consumable_margin = self.range_from_label(pvr_sheet, pvr_sheet_row, "Consumable Margin"))
            else:
                product_variables_ranges = pvr.ProductVariablesRanges.market_of(
                    name = pvr_name,
                    type = pvr_type,
                    existing_instance = product_variables_ranges,
                    unit_cost = self.range_from_label(pvr_sheet, pvr_sheet_row, "Unit Cost"),
                    unit_margin = self.range_from_label(pvr_sheet, pvr_sheet_row, "Unit Margin"),
                    sga_factor = self.range_from_label(pvr_sheet, pvr_sheet_row, "SG&A"),
                    yearly_unit_sales = self.range_from_label(pvr_sheet, pvr_sheet_row, "Yearly Unit Sales"),
                    yearly_unit_consumable_sales = self.range_from_label(pvr_sheet, pvr_sheet_row, "Yearly Unit Consumable Sales"),
                    years_of_consumable_sales = self.range_from_label(pvr_sheet, pvr_sheet_row, "Years of Consumable Sales"),
                    consumable_margin = self.range_from_label(pvr_sheet, pvr_sheet_row, "Consumable Margin"))
            
            if( pvr_exclude == None):
                mix_variables_ranges.append(product_variables_ranges)

        return company_constants, mix_variables_ranges

    def insert_plot_into_excel(excel_file_path_in, excel_file_path_out, image_path):
        # Load the workbook and select the sheet
        workbook = load_workbook(excel_file_path_in)
        sheet = workbook['Product Variables']

        # Load the image
        img = Image(image_path)
        
        # Insert the image into the sheet at the specified cell
        sheet.add_image(img, 'A16')

        # Save the workbook to the new file
        workbook.save(excel_file_path_out)
