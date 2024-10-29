import os
import tempfile
import sys
from enum import Enum
import numpy as np # linear algebra library
import matplotlib.pyplot as plt # plotting library
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

# Present Value (PV) function
def pv(future_value, rate, periods):
    return future_value / (1 + rate) ** periods

# Future Value (FV) function
def fv(present_value, rate, periods):
    return present_value * (1 + rate) ** periods

# Add a value to a given index in the array, growing the array if needed
def add_value_to_index(array, index, value):
    # Extend the array if the index is out of bounds
    if index >= len(array):
        array.extend([0] * (index - len(array) + 1))
    
    # Add the value to the specified index
    array[index] += value

# Company Constants (user input)
class CompanyConstants:
    def __init__(self,
                 market_return = 0.0, #What we would expect to earn on an investment in a financial market with a similar risk
                 yearly_development_fte_cost_pv = 17000, # How much it costs per person to run a product development team (in present dollars)
                 maximum_development_ftes = 6, # How many product developers are available to staff projects
                 development_cost_trend = 0.0, # How much development costs increase each year
                 product_cost_trend = 0.0, # How much product costs increase each year
                 product_price_trend = 0.0): # How much product prices increase each year

        self.market_return = market_return
        self.yearly_development_fte_cost_pv = yearly_development_fte_cost_pv
        self.maximum_development_ftes = maximum_development_ftes
        self.development_cost_trend = development_cost_trend
        self.product_cost_trend = product_cost_trend
        self.product_price_trend = product_price_trend

# Return the value of a range, or a single value if it is not a range
def safe_index(range, i):
    if isinstance(range, (int, float)):
        return range
    return range[i]

# Relationship between margin and cost factor
def cost_factor(margin):
    return 1 / (1 - margin)

# Product Variables Ranges (user input)
class ProductVariablesRanges:
    def __init__(self,
                 name = "",
                 type = "",
                 years_of_development_growth = 0, # Years of development FTE increase from zero (linear)
                 years_of_development_maturity = 0, # Years of development FTE full usage
                 years_of_development_decline = 0, #Years of development FTE decrease to zero (linear)
                 years_of_pilot = 0, # Years after development before sales begin
                 years_of_sales_growth = 0, # Years of sales increase from zero (linear)
                 years_of_sales_maturity = 0, # Years of full sales
                 years_of_sales_decline = 0, # Years of sales decrease to zero (linear)
                 development_ftes = 0, # The number of people (full time equivalents) on the development team
                 maintenance_ftes = 0, # The number of people that need to keep working on the product after development
                 years_of_maintenance = 0, # The number of years people need to keep working on the product after development
                 unit_cost_pv = 0, # The cost to manufacture one unit of the product (in present dollars)
                 unit_margin = 0, # The percentage of the selling price that exceeds cost of goods
                 sga_factor = 0, # The percentage of the selling price that is allocated to cover selling, general, and administrative
                 yearly_unit_sales = 0, # The number of units that are expected to be sold per year
                 yearly_unit_consumable_sales = 0, # The consumable sales per unit sold per year
                 years_of_consumable_sales = 0, # The number of years consumables will be sold following a unit sale
                 consumable_margin = 0): # The percentage of consumable sales that exceeds the consumable cost of goods

        self.name = name
        self.type = type
        self.years_mix_delay = 0
        self.years_of_development_growth = years_of_development_growth if years_of_development_growth is not None else [0, 0, 0]
        self.years_of_development_maturity = years_of_development_maturity if years_of_development_maturity is not None else [0, 0, 0]
        self.years_of_development_decline = years_of_development_decline if years_of_development_decline is not None else [0, 0, 0]
        self.years_of_pilot = years_of_pilot if years_of_pilot is not None else [0, 0, 0]
        self.years_of_sales_growth = years_of_sales_growth if years_of_sales_growth is not None else [0, 0, 0]
        self.years_of_sales_maturity = years_of_sales_maturity if years_of_sales_maturity is not None else [0, 0, 0]
        self.years_of_sales_decline = years_of_sales_decline if years_of_sales_decline is not None else [0, 0, 0]
        self.development_ftes = development_ftes if development_ftes is not None else [0, 0, 0]
        self.maintenance_ftes = maintenance_ftes if maintenance_ftes is not None else [0, 0, 0]
        self.years_of_maintenance = years_of_maintenance if years_of_maintenance is not None else [0, 0, 0]
        self.unit_cost_pv = unit_cost_pv if unit_cost_pv is not None else [0, 0, 0]
        self.unit_margin = unit_margin if unit_margin is not None else [0, 0, 0]
        self.sga_factor = sga_factor if sga_factor is not None else [0, 0, 0]
        self.yearly_unit_sales = yearly_unit_sales if yearly_unit_sales is not None else [0, 0, 0]
        self.yearly_unit_consumable_sales = yearly_unit_consumable_sales if yearly_unit_consumable_sales is not None else [0, 0, 0]
        self.years_of_consumable_sales = years_of_consumable_sales if years_of_consumable_sales is not None else [0, 0, 0]
        self.consumable_margin = consumable_margin if consumable_margin is not None else [0, 0, 0]
    
    @classmethod
    def market_of(
        cls,
        name,
        type,
        existing_instance,
        unit_cost,
        unit_margin,
        sga_factor,
        yearly_unit_sales,
        yearly_unit_consumable_sales,
        years_of_consumable_sales,
        consumable_margin):

        return cls(
            name = name,
            type = type,
            years_of_development_growth = 0,
            years_of_development_maturity = 0,
            years_of_development_decline = 0,
            years_of_pilot = 0,
            years_of_sales_growth = existing_instance.years_of_sales_growth,
            years_of_sales_maturity = existing_instance.years_of_sales_maturity,
            years_of_sales_decline = existing_instance.years_of_sales_decline,
            development_ftes = 0,
            maintenance_ftes = 0,
            years_of_maintenance = 0,
            unit_cost_pv = unit_cost if unit_cost is not None else existing_instance.unit_cost_pv,
            unit_margin = unit_margin if unit_margin is not None else existing_instance.unit_margin,
            sga_factor = sga_factor if sga_factor is not None else existing_instance.sga_factor,
            yearly_unit_sales = yearly_unit_sales if yearly_unit_sales is not None else existing_instance.yearly_unit_sales,
            yearly_unit_consumable_sales = yearly_unit_consumable_sales if yearly_unit_consumable_sales is not None else existing_instance.yearly_unit_consumable_sales,
            years_of_consumable_sales = years_of_consumable_sales if years_of_consumable_sales is not None else existing_instance.years_of_consumable_sales,
            consumable_margin = consumable_margin if consumable_margin is not None else existing_instance.consumable_margin)
    
# Return a single random number, given a low, expected, and high range, using a triangular distribution
# Just return the expected number if requested, or if the range is invalid
def triangle(a, just_expected = False):

    # if a is a single number, return it
    if isinstance(a, (int, float)):
        return a

    # if the range is invalid, return the likely (middle) value
    if(a[0] > a[1] or a[1] > a[2] or a[0] >= a[2]):
        return a[1]
    
    # if requested, just return the likely (middle) value
    if (just_expected):
        return a[1]
    
    # return a random number, given a low, likely, and high range, using a triangular distribution
    return np.random.triangular(a[0], a[1], a[2])

# Used to lock in all but one variable when computing a tornado sensitivity analysis
class Tornado(Enum):
    OFF = 0
    Dev_Ftes = 1
    Dev_Years = 2
    Maint_Ftes = 3
    Sales_Years = 4
    Unit_Cost = 5
    Margin = 6
    Yearly_Sales = 7

# Create a snapshot of the product variables with random values
class ProductVariablesSnapshot:
    def __init__(self, product_variables_ranges, tornado = Tornado.OFF):
        
        # convert various ranges to actual values using a triangular distribution (or use the likely value if a tornado sensitivity analysis is being performed)
        self.years_mix_delay = 0
        self.name = product_variables_ranges.name
        self.type = product_variables_ranges.type        
        self.development_ftes = triangle(product_variables_ranges.development_ftes, tornado != Tornado.OFF and tornado != Tornado.Dev_Ftes)
        self.years_of_development_growth = triangle(product_variables_ranges.years_of_development_growth)
        self.years_of_development_maturity = triangle(product_variables_ranges.years_of_development_maturity, tornado != Tornado.OFF and tornado != Tornado.Dev_Years)
        self.years_of_development_decline = triangle(product_variables_ranges.years_of_development_decline)
        self.maintenance_ftes = triangle(product_variables_ranges.maintenance_ftes, tornado != Tornado.OFF and tornado != Tornado.Maint_Ftes)
        self.years_of_maintenance = triangle(product_variables_ranges.years_of_maintenance)
        self.years_of_pilot = triangle(product_variables_ranges.years_of_pilot)
        self.years_of_sales_growth = triangle(product_variables_ranges.years_of_sales_growth)
        self.years_of_sales_maturity = triangle(product_variables_ranges.years_of_sales_maturity, tornado != Tornado.OFF and tornado != Tornado.Sales_Years)
        self.years_of_sales_decline = triangle(product_variables_ranges.years_of_sales_decline)
        self.unit_cost_pv = triangle(product_variables_ranges.unit_cost_pv, tornado != Tornado.OFF and tornado != Tornado.Unit_Cost)
        self.unit_margin = triangle(product_variables_ranges.unit_margin, tornado != Tornado.OFF and tornado != Tornado.Margin)
        self.sga_factor = triangle(product_variables_ranges.sga_factor)
        self.yearly_unit_sales = triangle(product_variables_ranges.yearly_unit_sales, tornado != Tornado.OFF and tornado != Tornado.Yearly_Sales)

        # compute the unit price
        self.unit_price_pv = self.unit_cost_pv * cost_factor(self.unit_margin)
        
        # precalculate the development full time equivalents for each month
        self.ftes_by_month = []
        for month in range (round(self.years_of_development_growth * 12)):
            self.ftes_by_month.append(self.development_ftes * month / (self.years_of_development_growth * 12))
        for month in range (round(self.years_of_development_maturity * 12)):
            self.ftes_by_month.append(self.development_ftes)
        for month in range (round(self.years_of_development_decline * 12)):
            self.ftes_by_month.append(self.development_ftes * (1 - month / (self.years_of_development_decline * 12)))
        for month in range (round(self.years_of_maintenance * 12)):
            self.ftes_by_month.append(self.maintenance_ftes)
        if (len(self.ftes_by_month) == 0):
            self.ftes_by_month.append(0)

        # precalculate the unit sales for a given month
        self.unit_sales_by_month = []
        for month in range (round(self.years_of_sales_growth * 12)):
            self.unit_sales_by_month.append(self.yearly_unit_sales * month / (self.years_of_sales_growth * 12) / 12)
        for month in range (round(self.years_of_sales_maturity * 12)):
            self.unit_sales_by_month.append(self.yearly_unit_sales / 12)
        for month in range (round(self.years_of_sales_decline * 12)):
            self.unit_sales_by_month.append(self.yearly_unit_sales * (1 - month / (self.years_of_sales_decline * 12)) / 12)
        if (len(self.unit_sales_by_month) == 0):
            self.unit_sales_by_month.append(0)

    # compute the delay before sales begin
    def years_before_sales(self):
        return self.years_mix_delay + self.years_of_development_growth + self.years_of_development_maturity + self.years_of_development_decline + self.years_of_pilot

    # compute the total number of years for the product
    def total_years(self):
        return self.years_before_sales() + self.years_of_sales_growth + self.years_of_sales_maturity + self.years_of_sales_decline

    # compute the development full time equivalents for a given month
    def development_ftes_this_mix_month(self, month):
        month -= self.years_mix_delay * 12
        month = round(month)
        return self.ftes_by_month[month] if 0 <= month < len(self.ftes_by_month) else 0
    
    # compute the unit sales for a given month
    def unit_sales_this_mix_month(self, month):
        month -= self.years_before_sales() * 12
        month = round(month)
        return self.unit_sales_by_month[month] if 0 <= month < len(self.unit_sales_by_month) else 0

# Create a snapshot of the mix of product variables with random values
class MixVariablesSnapshot:
    def __init__(self, mix_variables_ranges, tornado = Tornado.OFF):
        self.mix_variables_ranges = mix_variables_ranges
        self.mix_variables_snapshots = []
        for product_variables_ranges in mix_variables_ranges:
            product_variables_snapshot = ProductVariablesSnapshot(product_variables_ranges, tornado)
            self.mix_variables_snapshots.append(product_variables_snapshot)

# The result of a single NPV calculation
class NpvCalculationResult:
    def __init__(self):
        self.development_cost = 0
        self.sales = 0
        self.cost_of_goods = 0
        self.sga = 0
        self.unit_sales = 0
        self.ftes_by_month = []
        self.sales_by_month = []

    def npv(self):
        return self.sales - self.cost_of_goods - self.sga - self.development_cost

    def ros(self):
        if( self.sales == 0):
            return 0
        return self.npv() / self.sales

    def roi(self):
        if( self.development_cost == 0):
            return 0
        return self.npv() / self.development_cost

    def annualized_roi(self, years):
        if( years == 0):
            return 0
        return (1 + self.roi()) ** (1 / years) - 1
    
    def record_ftes(self, month, ftes):
        add_value_to_index(self.ftes_by_month, month, ftes)
    
    def record_sales(self, month, sales):
        add_value_to_index(self.sales_by_month, month, sales)
    
    def add(self, result):
        self.development_cost += result.development_cost
        self.sales += result.sales
        self.cost_of_goods += result.cost_of_goods
        self.sga += result.sga
        self.unit_sales += result.unit_sales
        for month in range(len(result.ftes_by_month)):
            add_value_to_index(self.ftes_by_month, month, result.ftes_by_month[month])
        for month in range(len(result.sales_by_month)):
            add_value_to_index(self.sales_by_month, month, result.sales_by_month[month])

# Calculate the NPV of a product
def calculate_product_npv(product_variables_snapshot, company_constants):

    # this is what we will calculate and return 
    product_result = NpvCalculationResult()

    # loop through all the months
    for month in range(round(product_variables_snapshot.total_years() * 12) + 1):
        mix_month = month + round(product_variables_snapshot.years_mix_delay * 12)
    
        # compute the development_ftes for this month
        development_ftes = product_variables_snapshot.development_ftes_this_mix_month(mix_month)

        # compute the unit sales for this month
        unit_sales = product_variables_snapshot.unit_sales_this_mix_month(mix_month)

        # compute the future value of the cost of development, unit cost, unit price, and sg&a for this month
        monthly_development_cost_fte_fv = fv(company_constants.yearly_development_fte_cost_pv / 12, company_constants.development_cost_trend / 12, mix_month)
        unit_cost_fv = fv(product_variables_snapshot.unit_cost_pv, company_constants.product_cost_trend / 12, mix_month)
        unit_price_fv = fv(product_variables_snapshot.unit_price_pv, company_constants.product_price_trend / 12, mix_month)
        sga_variable_fv = unit_price_fv * product_variables_snapshot.sga_factor

        # return everything to present value
        development_cost_pv = pv(development_ftes * monthly_development_cost_fte_fv, company_constants.market_return / 12, mix_month)
        sales_pv = pv(unit_sales * unit_price_fv, company_constants.market_return / 12, mix_month)
        cost_of_goods_pv = pv(unit_sales * unit_cost_fv, company_constants.market_return / 12, mix_month)
        sga_pv = pv(unit_sales * sga_variable_fv, company_constants.market_return / 12, mix_month)

        # add the results for this month to the total
        product_result.development_cost += development_cost_pv
        product_result.sales += sales_pv
        product_result.cost_of_goods += cost_of_goods_pv
        product_result.sga += sga_pv
        product_result.unit_sales += unit_sales
        product_result.record_ftes(mix_month, development_ftes)
        product_result.record_sales(mix_month, sales_pv)

    return product_result

# Calculate the years mix delay based on the maximum development FTEs
def calculate_years_mix_delay(maximum_development_ftes, minimum_years_mix_delay, allocated_ftes_by_month, new_ftes_by_month):
    months_mix_delay = round(minimum_years_mix_delay * 12)
    maximum_development_ftes = max(maximum_development_ftes, max(new_ftes_by_month))
    allocated_ftes_by_month.extend([0] * len(new_ftes_by_month))
    while(True):
        overallocated = False
        for month in range(len(new_ftes_by_month)):
            if allocated_ftes_by_month[month + months_mix_delay] + new_ftes_by_month[month] > maximum_development_ftes:
                overallocated = True
                break
        if not overallocated:
            return months_mix_delay / 12
        months_mix_delay += 1

# Calculate the NPV of a product mix
def calculate_mix_npv(mix_variables_snapshot, company_constants):
    mix_result = NpvCalculationResult()
    minimum_years_mix_delay = 0
    for product_variables_snapshot in mix_variables_snapshot.mix_variables_snapshots:
        if( product_variables_snapshot.type == "Product"):
            minimum_years_mix_delay = 0
        product_variables_snapshot.years_mix_delay = calculate_years_mix_delay(company_constants.maximum_development_ftes, minimum_years_mix_delay, mix_result.ftes_by_month, product_variables_snapshot.ftes_by_month)
        product_result = calculate_product_npv(product_variables_snapshot, company_constants)
        mix_result.add(product_result)
        minimum_years_mix_delay = product_variables_snapshot.years_before_sales()

    return mix_result

# Track all the results of multiple calculations
class SimulationTracker:
    def __init__(self):
        self.npvs_millions = []
        self.development_costs_millions = []
        self.unit_sales = []
        self.sales_millions = []
        self.ros = []
        self.roi = []
        self.ftes_by_month = []
        self.sales_by_month = []
    
    def add(self, result):
        self.npvs_millions.append(result.npv() / 1000000)
        self.development_costs_millions.append(result.development_cost / 1000000)
        self.unit_sales.append(result.unit_sales)
        self.sales_millions.append(result.sales / 1000000)
        self.ros.append(result.ros() * 100)
        self.roi.append(result.roi() * 100)
        for month in range(len(result.ftes_by_month)):
            add_value_to_index(self.ftes_by_month, month, result.ftes_by_month[month])
        for month in range(len(result.sales_by_month)):
            add_value_to_index(self.sales_by_month, month, result.sales_by_month[month])

class TornadoTracker:
    def __init__(self, tornado, name):
        self.tornado = tornado
        self.name = name
        self.min_value = float('inf')
        self.max_value = float('-inf')

    def add(self, value):
        if value < self.min_value:
            self.min_value = value
        if value > self.max_value:
            self.max_value = value
    
    def range(self):
        return self.max_value - self.min_value

class MonteCarloAnalyzer:
    def __init__(self, company_constants, mix_variables_ranges):
        self.company_constants = company_constants
        self.mix_variables_ranges = mix_variables_ranges
        self.simulation_tracker = SimulationTracker()
        self.tornado_trackers = []
        self.tornado_trackers.append(TornadoTracker(Tornado.Dev_Ftes, 'Dev FTEs'))
        self.tornado_trackers.append(TornadoTracker(Tornado.Dev_Years, 'Dev Years'))
        self.tornado_trackers.append(TornadoTracker(Tornado.Maint_Ftes, 'Maint FTEs'))
        self.tornado_trackers.append(TornadoTracker(Tornado.Sales_Years, 'Sales Years'))
        self.tornado_trackers.append(TornadoTracker(Tornado.Unit_Cost, 'Unit Cost'))
        self.tornado_trackers.append(TornadoTracker(Tornado.Margin, 'Margin'))
        self.tornado_trackers.append(TornadoTracker(Tornado.Yearly_Sales, 'Yearly Sales'))

    def analyze(self):
        # compute the monte carlo analysis
        simulations = 4000
        for i in range(simulations):
            mix_variables_snapshot = MixVariablesSnapshot(self.mix_variables_ranges)
            result = calculate_mix_npv(mix_variables_snapshot, self.company_constants)
            self.simulation_tracker.add(result)

        # normalize FTEs by dividing each value by the number of simulations
        for month in range(len(self.simulation_tracker.ftes_by_month)):
            self.simulation_tracker.ftes_by_month[month] /= simulations

        # normalize sales by dividing each value by the number of simulations
        for month in range(len(self.simulation_tracker.sales_by_month)):
            self.simulation_tracker.sales_by_month[month] /= simulations

        # compute the tornado analysis
        for i in range(100):    
            for tornado_tracker in self.tornado_trackers:
                mix_variables_snapshot = MixVariablesSnapshot(self.mix_variables_ranges, tornado_tracker.tornado)
                result = calculate_mix_npv(mix_variables_snapshot, self.company_constants)
                tornado_tracker.add(result.npv() / 1000000)

        # Sort the tornado trackers by range
        self.tornado_trackers.sort(key=lambda x: x.range())
    
    # create a histogram
    def create_histogram(self, data, bins, xlabel, subplot_position, rows, cols, color = 'blue'):
        plt.subplot(rows, cols, subplot_position)
        plt.hist(data, bins=bins, edgecolor='black', color=color)
        plt.yticks([])
        plt.xlabel(xlabel)

    # create a tornado plot
    def create_tornado_plot(self, names, ranges, min_values, xlabel, subplot_position, rows, cols):
        plt.subplot(rows, cols, subplot_position)
        plt.barh(names, ranges, left=min_values)
        plt.xlabel(xlabel)
    
    # create a line chart
    def create_line_chart(self, data, xlabel, ylabel, subplot_position, rows, cols, color = 'blue'):
        plt.subplot(rows, cols, subplot_position)
        years = np.arange(len(data)) / 12
        plt.plot(years, data, color=color)
        plt.xlabel(xlabel)
        plt.ylabel(ylabel)

    # plot the results
    def plot(self, file_path = ""):
        rows = 3
        cols = 3
        plt.figure(figsize=(10, 5))

        self.create_histogram(self.simulation_tracker.unit_sales, 20, 'Sales (units)', 1, rows, cols, 'blue')
        self.create_histogram(self.simulation_tracker.sales_millions, 20, 'Sales ($ millions)', 2, rows, cols, 'red')
        self.create_histogram(self.simulation_tracker.development_costs_millions, 20, 'Development ($ millions)', 3, rows, cols, 'green')
        self.create_histogram(self.simulation_tracker.npvs_millions, 20, 'NPV ($ millions)', 4, rows, cols, 'purple')
        self.create_histogram(self.simulation_tracker.ros, 20, 'ROS (%)', 5, rows, cols, 'orange')
        self.create_histogram(self.simulation_tracker.roi, 20, 'ROI (%)', 6, rows, cols, 'pink')

        tornado_names = [tornado_tracker.name for tornado_tracker in self.tornado_trackers]
        tornado_ranges = [tornado_tracker.range() for tornado_tracker in self.tornado_trackers]
        tornado_min_values = [tornado_tracker.min_value for tornado_tracker in self.tornado_trackers]
        self.create_tornado_plot(tornado_names, tornado_ranges, tornado_min_values, 'NPV Sensitivity ($ millions)', 7, rows, cols)

        self.create_line_chart(self.simulation_tracker.ftes_by_month, 'Years', 'Ftes', 8, rows, cols, 'black')
        self.create_line_chart(self.simulation_tracker.sales_by_month, 'Years', 'Monthly Sales ($)', 9, rows, cols, 'black')

        plt.tight_layout()
        if( file_path != ""):
            plt.savefig(file_path)
        else:
            plt.show()
        plt.close()

def read_excel_data(file_path):
    # Load the workbook and select the sheets
    workbook = load_workbook(file_path, data_only=True)
    company_constants_sheet = workbook['Company Constants']

    # Initialize the CompanyConstants instance
    company_constants = CompanyConstants(
        market_return = company_constants_sheet.cell(row=2, column=2).value,
        yearly_development_fte_cost_pv = company_constants_sheet.cell(row=3, column=2).value,
        maximum_development_ftes = company_constants_sheet.cell(row=4, column=2).value,
        development_cost_trend = company_constants_sheet.cell(row=5, column=2).value,
        product_cost_trend = company_constants_sheet.cell(row=6, column=2).value,
        product_price_trend = company_constants_sheet.cell(row=7, column=2).value)

    def row_from_label(sheet, start_row, label):
        for row in range(start_row, sheet.max_row + 1):
            cell_value = sheet.cell(row=row, column=1).value
            if cell_value == label:
                return row
        return 0

    def range_from_label(sheet, start_row, label):
        for row in range(start_row, sheet.max_row + 1):
            cell_value = sheet.cell(row=row, column=1).value
            if cell_value == None:
                return None
            if cell_value == label:
                return [
                    sheet.cell(row=row, column=2).value,
                    sheet.cell(row=row, column=3).value,
                    sheet.cell(row=row, column=4).value
                ]

    mix_variables_ranges = []
    mix_sheet = workbook['Mix']
    product_variables_ranges = ProductVariablesRanges()

    for mix_sheet_row in range(2, mix_sheet.max_row + 1):
        pvr_sheet_name = mix_sheet.cell(row=mix_sheet_row, column=1).value
        if( pvr_sheet_name == None ):
            break
        pvr_name = mix_sheet.cell(row=mix_sheet_row, column=2).value
        pvr_type = mix_sheet.cell(row=mix_sheet_row, column=3).value
        pvr_exclude = mix_sheet.cell(row=mix_sheet_row, column=4).value
        pvr_sheet = workbook[pvr_sheet_name]
        pvr_sheet_row = row_from_label(pvr_sheet, 1, pvr_name)
        if( pvr_sheet_row == 0):
            continue

        if( pvr_type == "Product"):
            product_variables_ranges = ProductVariablesRanges(
                name = pvr_name,
                type = pvr_type,
                years_of_development_growth = range_from_label(pvr_sheet, pvr_sheet_row, "Years of Development Growth"),
                years_of_development_maturity = range_from_label(pvr_sheet, pvr_sheet_row, "Years of Development Maturity"),
                years_of_development_decline = range_from_label(pvr_sheet, pvr_sheet_row, "Years of Development Decline"),
                years_of_pilot = range_from_label(pvr_sheet, pvr_sheet_row, "Years of Pilot"),
                years_of_sales_growth = range_from_label(pvr_sheet, pvr_sheet_row, "Years of Sales Growth"),
                years_of_sales_maturity = range_from_label(pvr_sheet, pvr_sheet_row, "Years of Sales Maturity"),
                years_of_sales_decline = range_from_label(pvr_sheet, pvr_sheet_row, "Years of Sales Decline"),
                development_ftes = range_from_label(pvr_sheet, pvr_sheet_row, "Development FTEs"),
                maintenance_ftes = range_from_label(pvr_sheet, pvr_sheet_row, "Maintenance FTEs"),
                years_of_maintenance = range_from_label(pvr_sheet, pvr_sheet_row, "Years of Maintenance"),
                unit_cost_pv = range_from_label(pvr_sheet, pvr_sheet_row, "Unit Cost"),
                unit_margin = range_from_label(pvr_sheet, pvr_sheet_row, "Unit Margin"),
                sga_factor = range_from_label(pvr_sheet, pvr_sheet_row, "SG&A"),
                yearly_unit_sales = range_from_label(pvr_sheet, pvr_sheet_row, "Yearly Unit Sales"),
                yearly_unit_consumable_sales = range_from_label(pvr_sheet, pvr_sheet_row, "Yearly Unit Consumable Sales"),
                years_of_consumable_sales = range_from_label(pvr_sheet, pvr_sheet_row, "Years of Consumable Sales"),
                consumable_margin = range_from_label(pvr_sheet, pvr_sheet_row, "Consumable Margin"))
        else:
            product_variables_ranges = ProductVariablesRanges.market_of(
                name = pvr_name,
                type = pvr_type,
                existing_instance = product_variables_ranges,
                unit_cost = range_from_label(pvr_sheet, pvr_sheet_row, "Unit Cost"),
                unit_margin = range_from_label(pvr_sheet, pvr_sheet_row, "Unit Margin"),
                sga_factor = range_from_label(pvr_sheet, pvr_sheet_row, "SG&A"),
                yearly_unit_sales = range_from_label(pvr_sheet, pvr_sheet_row, "Yearly Unit Sales"),
                yearly_unit_consumable_sales = range_from_label(pvr_sheet, pvr_sheet_row, "Yearly Unit Consumable Sales"),
                years_of_consumable_sales = range_from_label(pvr_sheet, pvr_sheet_row, "Years of Consumable Sales"),
                consumable_margin = range_from_label(pvr_sheet, pvr_sheet_row, "Consumable Margin"))
        
        if( pvr_exclude == None):
            mix_variables_ranges.append(product_variables_ranges)

    return company_constants, mix_variables_ranges

def insert_plot_into_excel(excel_file_path_in, excel_file_path_out, image_path):
    # Load the workbook and select the sheet
    workbook = load_workbook(excel_file_path_in)
    sheet = workbook['Product Variables']

    # Load the image
    img = Image(image_path)
    
    # Insert the image into the sheet at the specified cell
    sheet.add_image(img, 'A16')

    # Save the workbook to the new file
    workbook.save(excel_file_path_out)

# Check if there are any command line arguments
if len(sys.argv) == 2:
    excel_file_path = sys.argv[1]
    plot_file_path = os.path.join(tempfile.gettempdir(), 'ppm.png')
    company_constants, mix_variables_ranges = read_excel_data(excel_file_path)
else:
    excel_file_path = ""
    plot_file_path = ""
    company_constants = CompanyConstants()
    product1 = ProductVariablesRanges()
    mix_variables_ranges = [product1]

# Run the Monte Carlo simulation
plot_file_path = ""
monte_carlo = MonteCarloAnalyzer(company_constants, mix_variables_ranges)
monte_carlo.analyze()
monte_carlo.plot(plot_file_path)

# Save the plot
if( plot_file_path != ""):
    insert_plot_into_excel(excel_file_path, excel_file_path, plot_file_path)